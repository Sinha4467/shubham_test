import openpyxl

path_file = openpyxl.load_workbook("F:/result project/Input_Result Sheet")
sheets = path_file.sheetnames
path_file.active

print("active sheet :", path_file.active.title)

rows = path_file.active.max_row
print(rows)

class student:

    def __init__ (self, cell_listj,cell_listk,cell_listl,cell_listm,cell_listn,cell_listo,cell_listp,cell_listq,cell_listr,cell_lists,cell_listt,cell_listu,cell_listv,cell_listw,cell_listx, count1,count2,count3,count4,count5,count6,count7,count8,count9,count10,count11,count12,count13,count14,count15):
        self.cell_listj =cell_listj
        self.cell_listk =cell_listk
        self.cell_listl =cell_listl
        self.cell_listm =cell_listm
        self.cell_listn =cell_listn
        self.cell_listo =cell_listo
        self.cell_listp =cell_listp
        self.cell_listq =cell_listq
        self.cell_listr =cell_listr
        self.cell_lists =cell_lists
        self.cell_listt =cell_listt
        self.cell_listu =cell_listu
        self.cell_listv =cell_listv
        self.cell_listw =cell_listw
        self.cell_listx =cell_listx
        self.count1=count1
        self.count2=count2
        self.count3=count3
        self.count4=count4
        self.count5=count5
        self.count6=count6
        self.count7=count7
        self.count8=count8
        self.count9=count9
        self.count10=count10
        self.count11=count11
        self.count12=count12
        self.count13=count13
        self.count14=count14
        self.count15=count15



    def cn(self):


        for i in range(3, rows+1):
            Cellj = path_file.active.cell(row = i, column = 10)
            self.cell_listj.append(Cellj.value)

        for i in range(rows-3):
            if self.cell_listj[i]<50:
                self.count1 += 1
        path_file.active.cell(row = 1, column = 10).value = self.count1

        for i in range(3, rows+1):
            Cellk = path_file.active.cell(row = i, column = 11)
            self.cell_listk.append(Cellk.value)

        for i in range(rows-3):
            if self.cell_listk[i]<50:
                self.count2 += 1
        path_file.active.cell(row = 1, column = 11).value = self.count2

        for i in range(3, rows+1):
            Celll = path_file.active.cell(row = i, column = 12)
            self.cell_listl.append(Celll.value)

        for i in range(rows-3):
            if self.cell_listl[i]<50:
                self.count3 += 1
        path_file.active.cell(row = 1, column = 12).value = self.count3

        for i in range(3, rows+1):
            Cellm = path_file.active.cell(row = i, column = 13)
            self.cell_listm.append(Cellm.value)

        for i in range(rows-3):
            if self.cell_listm[i]<50:
                self.count4 += 1
        path_file.active.cell(row = 1, column = 13).value = self.count4

        for i in range(3, rows+1):
            Celln = path_file.active.cell(row = i, column = 14)
            self.cell_listn.append(Celln.value)

        for i in range(rows-3):
            if self.cell_listn[i]<50:
                self.count5 += 1
        path_file.active.cell(row = 1, column = 14).value = self.count5

        for i in range(3, rows+1):
            Cello = path_file.active.cell(row = i, column = 15)
            self.cell_listo.append(Cello.value)

        for i in range(rows-3):
            if self.cell_listo[i]<50:
                self.count6 += 1
        path_file.active.cell(row = 1, column = 15).value = self.count6

        for i in range(3, rows+1):
            Cellp = path_file.active.cell(row = i, column = 16)
            self.cell_listp.append(Cellp.value)

        for i in range(rows-3):
            if self.cell_listp[i]<50:
                self.count7 += 1
        path_file.active.cell(row = 1, column = 16).value = self.count7

        for i in range(3, rows+1):
            Cellq = path_file.active.cell(row = i, column = 17)
            self.cell_listq.append(Cellq.value)

        for i in range(rows-3):
            if self.cell_listq[i]<50:
                self.count8 += 1
        path_file.active.cell(row = 1, column = 17).value = self.count8

        for i in range(3, rows+1):
            Cellr = path_file.active.cell(row = i, column = 18)
            self.cell_listr.append(Cellr.value)

        for i in range(rows-3):
            if self.cell_listr[i]<50:
                self.count9 += 1
        path_file.active.cell(row = 1, column = 18).value = self.count9

        for i in range(3, rows+1):
            Cells = path_file.active.cell(row = i, column = 19)
            self.cell_lists.append(Cells.value)

        for i in range(rows-3):
            if self.cell_lists[i]<50:
                self.count10 += 1
        path_file.active.cell(row = 1, column = 19).value = self.count10

        for i in range(3, rows+1):
            Cellt = path_file.active.cell(row = i, column = 20)
            self.cell_listt.append(Cellt.value)

        for i in range(rows-3):
            if self.cell_listt[i]<50:
                self.count11 += 1
        path_file.active.cell(row = 1, column = 20).value = self.count11

        for i in range(3, rows+1):
            Cellu = path_file.active.cell(row = i, column = 21)
            self.cell_listu.append(Cellu.value)

        for i in range(rows-3):
            if self.cell_listu[i]<50:
                self.count12 += 1
        path_file.active.cell(row = 1, column = 21).value = self.count12

        for i in range(3, rows+1):
            Cellv = path_file.active.cell(row = i, column = 22)
            self.cell_listv.append(Cellv.value)

        for i in range(rows-3):
            if self.cell_listv[i]<50:
                self.count13 += 1
        path_file.active.cell(row = 1, column = 22).value = self.count13

        for i in range(3, rows+1):
            Cellw = path_file.active.cell(row = i, column = 23)
            self.cell_listw.append(Cellw.value)

        for i in range(rows-3):
            if self.cell_listw[i]<50:
                self.count14 += 1

        path_file.active.cell(row = 1, column = 23).value = self.count14

        for i in range(3, rows+1):
            Cellx = path_file.active.cell(row = i, column = 24)
            self.cell_listx.append(Cellx.value)

        for i in range(rows-3):
            if self.cell_listx[i]<50:
                self.count15 += 1

        path_file.active.cell(row = 1, column = 24).value = self.count15


        def check(ls3, val, len_rows):
            z =[]
            for i in range(len_rows):
                if val>ls3[i]:
                    z.append(i)
                else:
                    continue
            return len(z)

        for i in range(rows-2):
            ls3 =[self.cell_listj[i],self.cell_listk[i],self.cell_listl[i],self.cell_listm[i],self.cell_listn[i],self.cell_listo[i],self.cell_listp[i],self.cell_listq[i],self.cell_listr[i],self.cell_lists[i],self.cell_listt[i],self.cell_listu[i],self.cell_listv[i],self.cell_listw[i],self.cell_listx[i]]
            len_rows=len(ls3)
            kate = check(ls3,50,len_rows)
            path_file.active.cell(row = i+3, column = 9).value = kate

        for i in range(rows-2):
            Total = self.cell_listj[i]+self.cell_listk[i]+self.cell_listl[i]+self.cell_listm[i]+self.cell_listn[i]+self.cell_listo[i]+self.cell_listp[i]+self.cell_listq[i]+self.cell_listr[i]+self.cell_lists[i]+self.cell_listt[i]+self.cell_listu[i]+self.cell_listv[i]+self.cell_listw[i]+self.cell_listx[i]
            Percentage = round((Total/1500)*100, 2)

            len_rows=len(ls3)
            path_file.active.cell(row = i+3, column = 25).value = Total
            path_file.active.cell(row = i+3, column = 26).value = Percentage

        path_file.save("F:/result project/Output_Result Sheet")





students = list()
for i in range(rows):
    students.append(student([],[],[],[],[],[],[],[],[],[],[],[],[],[],[], 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0))

students[0].cn()



